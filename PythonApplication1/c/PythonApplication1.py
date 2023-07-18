import array
from ast import And
from statistics import geometric_mean
from tkinter import *
from tkinter.filedialog import askopenfilename, asksaveasfile
from tkinter.font import BOLD
from turtle import bgcolor, left
import openpyxl
from openpyxl import load_workbook
import tkinter.filedialog as fd
import datetime
from tkinter import ttk

class Application(Frame):

    def genUVS(self):
        wb = load_workbook(self.filename, data_only=True)
        ws1=wb["ВС_Конф"]
        Name=""
        maxstr=52
        arrayTemp=[0 for j in range(0,13)]
        self.strArray=[]

        #Gen Config_UVS
        self.strArray.append('VAR_GLOBAL CONSTANT')
        self.strArray.append('\n	Config_UVS : ARRAY [1..50] OF sConfig_UVS := [')


        for i in range(2,maxstr):
            Name = str(ws1.cell(row=i, column=2).value)

            if Name == 'None':
                Name="Резерв"
                for j in range(2,13):
                    arrayTemp[j]="0"
            else:
                for j in range(2,13):
                    arrayTemp[j] = str(ws1.cell(row=i, column=j+1).value)

                for j in range(0,13):
                    if arrayTemp[j] == 'None':
                        arrayTemp[j]="0"

            if (maxstr-1)==i:
                self.strArray.append("\n    (NUM_GRP:="+arrayTemp[2]+", PC_USE:="+arrayTemp[3]+", nMP:="+arrayTemp[4]+", nPC_type:="+arrayTemp[5]+", nPC:="+arrayTemp[6]+
                                     ", nEC:="+arrayTemp[7]+", nSecEC:="+arrayTemp[8]+", nAvar:="+arrayTemp[9]+", nCV:="+arrayTemp[10]+", nCMD_ABB:="+arrayTemp[11]+
                                     ", nCMD_ABO:="+arrayTemp[12]+")]; (*"+" "+str(i-1)+"  "+Name+"*)")
                self.strArray.append("\nEND_VAR")
            else:
                self.strArray.append("\n    (NUM_GRP:="+arrayTemp[2]+", PC_USE:="+arrayTemp[3]+", nMP:="+arrayTemp[4]+", nPC_type:="+arrayTemp[5]+", nPC:="+arrayTemp[6]+
                                     ", nEC:="+arrayTemp[7]+", nSecEC:="+arrayTemp[8]+", nAvar:="+arrayTemp[9]+", nCV:="+arrayTemp[10]+", nCMD_ABB:="+arrayTemp[11]+
                                     ", nCMD_ABO:="+arrayTemp[12]+"), (*"+" "+str(i-1)+"  "+Name+"*)")
        self.savefile()




    def genKTPRS(self):
        wb = load_workbook(self.filename, data_only=True)
        ws1=wb["Пред. Параметры"]
        maxstr=302
        arrayTemp=[[str(0) for j in range(7)] for i in range(maxstr)]
        self.strArray=[]


        for i in range(2,maxstr):
            Name = str(ws1.cell(row=i, column=1).value)
            if Name == 'None':
                Name="Резерв"
                arrayTemp[i][1]="Резерв"
                for j in range(2,7):
                    arrayTemp[i][j]="0"
            elif Name == 'Резерв':
                for j in range(2,7):
                    if arrayTemp[i][j]== 'None':
                        arrayTemp[i][2]= '0'
            else:
                for j in range(1,7):
                    arrayTemp[i][j]=str(ws1.cell(row=i, column=j+1).value)
                    if arrayTemp[i][j]=='None':
                        arrayTemp[i][j]='0'
        #Gen Config_KTPRS
        self.strArray.append('\n VAR_GLOBAL CONSTANT')
        self.strArray.append('\n Config_KTPRS :  ARRAY[1..300] OF sConfig_Signaliz:=[')
        for i in range(2,maxstr):
            if (maxstr-1)==i:
                end = ')];'
            else:
                end = '),'

            PointType=arrayTemp[i][3]
            if PointType=='ДП':
                PointType = '0'
            if PointType=='None':
                PointType = '0'
            elif PointType=='ИП':
                PointType = '1'
            elif PointType=='ИП_КЗ':
                PointType = '2'

            self.strArray.append("\n (No:="+arrayTemp[i][4]+", PointType:="+PointType+", Inversiya:="+arrayTemp[i][6]+", SubType:="+arrayTemp[i][5]+end+" (*"+" "+str(i-1)+"  "+arrayTemp[i][1]+"*)")
        self.strArray.append("\n END_VAR")
        self.savefile()

    def genUTS(self):
        wb = load_workbook(self.filename, data_only=True)
        ws1=wb["Табло и сирены"]
        maxstr=302
        arrayTemp=[[str(0) for j in range(7)] for i in range(maxstr)]
        self.strArray=[]

        for i in range(2,maxstr):
            Name = str(ws1.cell(row=i, column=2).value)
            if Name == 'None':
                Name="Резерв"
                arrayTemp[i][1]="Резерв"
                for j in range(2,6):
                    arrayTemp[i][j]="0"
            elif Name == 'Резерв':
                for j in range(2,7):
                    if arrayTemp[i][j]== 'None':
                        arrayTemp[i][2]= '0'
            else:
                for j in range(0,6):
                    arrayTemp[i][j]=str(ws1.cell(row=i, column=j+1).value)
                    if arrayTemp[i][j]=='None':
                        arrayTemp[i][j]='0'

        #Gen Config_Alarm
        self.strArray.append('\n VAR_GLOBAL CONSTANT')
        self.strArray.append('\n Config_Alarm :  ARRAY[1..300] OF sConfig_Alarm:=[')
        for i in range(2,maxstr):
            if (maxstr-1)==i:
                end = ')];'
            else:
                end = '),'
            self.strArray.append("\n (Enable:="+arrayTemp[i][4]+", nDO:="+arrayTemp[i][2]+", nCorrCV:="+arrayTemp[i][3]+", isSiren:="+arrayTemp[i][5]+", EnableCVCheck:="+arrayTemp[i][6]+end+" (*"+" "+str(i-1)+"  "+arrayTemp[i][1]+"*)")
        self.strArray.append("\n END_VAR")
        self.savefile()

    def genIP(self):
        wb = load_workbook(self.filename, data_only=True)
        ws1=wb["ИП"]
        ws2=wb["ИП_КЗ_Об"]
        maxstr=102
        maxstrKz_Obriv=100
        arrayIP=[[str(0) for j in range(29)] for i in range(maxstr)]
        arrayKz_Obriv=[[str(0) for j in range(3)] for i in range(maxstr)]

        self.strArray=[]

        for i in range(2,maxstr):
            Name = str(ws1.cell(row=i, column=2).value)

            if Name == 'None':
                Name="Резерв"
                for j in range(0,29):
                    arrayIP[i][j]="0"
            elif Name == 'Резерв':
                for j in range(0,29):
                    if arrayIP[i][j]== 'None':
                        arrayIP[i][j]= '0'
            else:
                for j in range(0,29):
                    arrayIP[i][j]=str(ws1.cell(row=i, column=j+1).value)

        for i in range(2,maxstrKz_Obriv):
            Name = str(ws2.cell(row=i, column=2).value)
            if Name == 'None':
                Name="Резерв"
                arrayKz_Obriv[i][1]="Резерв"
                arrayKz_Obriv[i][2]="0"
            elif Name == 'Резерв':
                if arrayKz_Obriv[i][2]== 'None':
                    arrayKz_Obriv[i][2]= '0'
            else:
                for j in range(0,3):
                    arrayKz_Obriv[i][j]=str(ws2.cell(row=i, column=j+1).value)

        #Gen ElAllowedRange
        self.strArray.append('\n	--------------Copy to sec OIP init ------------------')
        self.strArray.append('\nIF NOT InitOIP THEN ')
        self.strArray.append('\n    InitOIP:=TRUE;')
        self.strArray.append('\n')
        for i in range(2,maxstr):
            self.strArray.append("\n    Config_ZI["+str(i-1)+"].ElAllowedRange.Bottom:="+arrayIP[i][10]+"; Config_ZI["+str(i-1)+"].ElAllowedRange.Top:="+arrayIP[i][11]+"; (*"+" "+str(i-1)+"  "+arrayIP[i][1]+"*)")

   
        #Gen ValRange
        self.strArray.append('\n')
        for i in range(2,maxstr):
            self.strArray.append("\n Config_ZI["+str(i-1)+"].ValRange.Bottom:="+arrayIP[i][8]+"; Config_ZI["+str(i-1)+"].ValRange.Top:="+arrayIP[i][9]+"; (*"+" "+str(i-1)+"  "+arrayIP[i][1]+"*)")

        #Gen ElRange
        self.strArray.append('\n')
        for i in range(2,maxstr):
            self.strArray.append("\n Config_ZI["+str(i-1)+"].ElRange.Bottom:="+arrayIP[i][6]+"; Config_ZI["+str(i-1)+"].ElRange.Top:="+arrayIP[i][7]+"; (*"+" "+str(i-1)+"  "+arrayIP[i][1]+"*)")

        #Gen Hist
        self.strArray.append('\n')
        for i in range(2,maxstr):
            self.strArray.append("\n Config_ZI["+str(i-1)+"].Hist:="+arrayIP[i][12]+"; (*"+" "+str(i-1)+"  "+arrayIP[i][1]+"*)")
        
        self.strArray.append('\nEnd_IF;')
        self.strArray.append('\n\n')


        self.strArray.append('\n	--------------Copy to Const OIP------------------')
        #Gen Config_ZI_HRDW
        self.strArray.append('\n VAR_GLOBAL CONSTANT')
        self.strArray.append('\n Config_ZI_HRDW :  ARRAY[1..300] OF sConfig_ZI_HRDW:=[')
        for i in range(2,maxstr):
            if (maxstr-1)==i:
                end = ')];'
            else:
                end = '),'
            self.strArray.append("\n (NumMod:="+arrayIP[i][3]+", NumChan:="+arrayIP[i][4]+end+" (*"+" "+str(i-1)+"  "+arrayIP[i][1]+"*)")
        self.strArray.append('\n \n ')

        self.strArray.append('\n	Config_ZI_KZ_Obriv:  ARRAY[1..100] OF UINT:=[')
        for i in range(2,maxstr):
            if (maxstr-1)==i:
                end = '];'
            else:
                end = ','
            self.strArray.append("\n "+arrayKz_Obriv[i][2]+end+" (*"+" "+str(i-1)+"  "+arrayKz_Obriv[i][1]+"*)")

        self.strArray.append("\n END_VAR")
            
        self.savefile()



    def genZD(self):
        wb = load_workbook(self.filename, data_only=True)
        ws1=wb["Задвижки"]
        Name=""
        NumMode=""
        NumChannel=""
        maxstr=100
        arrayZDV=[0 for j in range(0,29)]
        self.strArray=[]

        #Gen Config_UZD
        self.strArray.append('VAR_GLOBAL CONSTANT')
        self.strArray.append('\n	Config_UZD : ARRAY [1..100] OF sConfig_UZD:=[')


        for i in range(2,maxstr):
            Name = str(ws1.cell(row=i, column=2).value)

            if Name == 'None':
                Name="Резерв"
                for j in range(0,29):
                    arrayZDV[j]="0"
            else:
                arrayZDV[0] = str(ws1.cell(row=i, column=11).value)
                arrayZDV[1] = str(ws1.cell(row=i, column=12).value)
                arrayZDV[2] = str(ws1.cell(row=i, column=13).value)
                arrayZDV[3] = str(ws1.cell(row=i, column=14).value)
                arrayZDV[4] = str(ws1.cell(row=i, column=15).value)
                arrayZDV[5] = str(ws1.cell(row=i, column=16).value)
                arrayZDV[6] = str(ws1.cell(row=i, column=17).value)
                arrayZDV[7] = str(ws1.cell(row=i, column=18).value)
                arrayZDV[8] = str(ws1.cell(row=i, column=19).value)
                arrayZDV[9] = str(ws1.cell(row=i, column=20).value)
                arrayZDV[10] = str(ws1.cell(row=i, column=21).value)
                arrayZDV[11] = str(ws1.cell(row=i, column=22).value)
                arrayZDV[12] = str(ws1.cell(row=i, column=23).value)
                arrayZDV[13] = str(ws1.cell(row=i, column=24).value)
                arrayZDV[14] = str(ws1.cell(row=i, column=25).value)
                arrayZDV[15] = str(ws1.cell(row=i, column=26).value)
                arrayZDV[16] = str(ws1.cell(row=i, column=27).value)
                arrayZDV[17] = str(ws1.cell(row=i, column=28).value)
                arrayZDV[18] = str(ws1.cell(row=i, column=29).value)
                arrayZDV[19] = str(ws1.cell(row=i, column=30).value)
                arrayZDV[20] = str(ws1.cell(row=i, column=31).value)
                arrayZDV[21] = str(ws1.cell(row=i, column=32).value)

                for j in range(0,29):
                    if arrayZDV[j] == 'None':
                        arrayZDV[j]="0"

            if (maxstr-1)==i:
                self.strArray.append("\n(nMPO:="+arrayZDV[0]+", nMPZ:="+arrayZDV[1]+", nKVO:="+arrayZDV[2]+", nKVZ:="+arrayZDV[3]+", nVMMO:="+arrayZDV[1]+
                                     ", nVMMZ:="+arrayZDV[1]+", nMuft:="+arrayZDV[1]+", nAvar:="+arrayZDV[1]+", nFault:="+arrayZDV[1]+", nDist:="+arrayZDV[1]+
                                     ", nNapr:="+arrayZDV[1]+", nNapr_Sec:="+arrayZDV[1]+", nCorrCO:="+arrayZDV[1]+", nCorrCZ:="+arrayZDV[1]+", nBRU_Dist:="+arrayZDV[1]+
                                     ", nBRU_Open:="+arrayZDV[1]+", nBRU_Close:="+arrayZDV[1]+", nBRU_Stop:="+arrayZDV[1]+", nRez:="+arrayZDV[1]+", nCMD_Open:="+arrayZDV[1]+
                                     ", nCMD_Close:="+arrayZDV[1]+", nCMD_Stop:="+arrayZDV[1]+", nCMD_Stop_Open:="+arrayZDV[1]+", nCMD_Stop_Close:="+arrayZDV[1]+")]; (*"+" "+str(i-1)+"  "+Name+"*)")
                #self.strArray.append("\nEND_VAR")
            else:
                self.strArray.append("\n(nMPO:="+arrayZDV[0]+", nMPZ:="+arrayZDV[1]+", nKVO:="+arrayZDV[2]+", nKVZ:="+arrayZDV[3]+", nVMMO:="+arrayZDV[1]+
                                     ", nVMMZ:="+arrayZDV[1]+", nMuft:="+arrayZDV[1]+", nAvar:="+arrayZDV[1]+", nFault:="+arrayZDV[1]+", nDist:="+arrayZDV[1]+
                                     ", nNapr:="+arrayZDV[1]+", nNapr_Sec:="+arrayZDV[1]+", nCorrCO:="+arrayZDV[1]+", nCorrCZ:="+arrayZDV[1]+", nBRU_Dist:="+arrayZDV[1]+
                                     ", nBRU_Open:="+arrayZDV[1]+", nBRU_Close:="+arrayZDV[1]+", nBRU_Stop:="+arrayZDV[1]+", nRez:="+arrayZDV[1]+", nCMD_Open:="+arrayZDV[1]+
                                     ", nCMD_Close:="+arrayZDV[1]+", nCMD_Stop:="+arrayZDV[1]+", nCMD_Stop_Open:="+arrayZDV[1]+", nCMD_Stop_Close:="+arrayZDV[1]+"), (*"+" "+str(i-1)+"  "+Name+"*)")


        #Gen UZD Type
        self.strArray.append('\n')
        self.strArray.append('\n		(*Конфигурация типа задвижки*)')
        self.strArray.append('\n		Config_UZD_type :  ARRAY[1..100] OF ZDCfg_struct:=[')

        for i in range(2,maxstr):
            Name = str(ws1.cell(row=i, column=2).value)

            if Name == 'None':
                Name="Резерв"
                for j in range(0,7):
                    arrayZDV[j]="0"
            else:
                arrayZDV[0] = str(ws1.cell(row=i, column=4).value)
                arrayZDV[1] = str(ws1.cell(row=i, column=5).value)
                arrayZDV[2] = str(ws1.cell(row=i, column=6).value)
                arrayZDV[3] = str(ws1.cell(row=i, column=7).value)
                arrayZDV[4] = str(ws1.cell(row=i, column=8).value)
                arrayZDV[5] = str(ws1.cell(row=i, column=9).value)
                arrayZDV[6] = str(ws1.cell(row=i, column=10).value)

                for j in range(0,7):
                    if arrayZDV[j] == 'None':
                        arrayZDV[j]="0"

            if (maxstr-1)==i:
                self.strArray.append("\n(slDist:="+arrayZDV[0]+", sl2Stop:="+arrayZDV[1]+", slBUR:="+arrayZDV[2]+", slCOz:="+arrayZDV[3]+", slCZz:="+arrayZDV[1]+
                                     ", slYesEC:="+arrayZDV[1]+", RS_OFF:="+arrayZDV[1]+")]; (*"+" "+str(i-1)+"  "+Name+"*)")
                self.strArray.append("\nEND_VAR")
            else:
                self.strArray.append("\n(slDist:="+arrayZDV[0]+", sl2Stop:="+arrayZDV[1]+", slBUR:="+arrayZDV[2]+", slCOz:="+arrayZDV[3]+", slCZz:="+arrayZDV[1]+
                                     ", slYesEC:="+arrayZDV[1]+", RS_OFF:="+arrayZDV[1]+"), (*"+" "+str(i-1)+"  "+Name+"*)")
        

        self.savefile()




    def genSignaliz(self):
        wb = load_workbook(self.filename)
        ws1=wb["Сигнализации"]
        Name=""
        NumMode=""
        NumChannel=""
        maxstr=1000
        self.strArray=[]

        self.strArray.append('VAR_GLOBAL CONSTANT')
        self.strArray.append('\nConfig_Signaliz :  ARRAY[1..1000] OF sConfig_Signaliz:=[')

        for i in range(2,maxstr):
            Name = str(ws1.cell(row=i, column=2).value)

            if Name == 'None':
                Name="Резерв"
                NumParam="0"
                TypeParam = "0"
                SubTypeParam="0"
                Inversia="0"
            else:
                NumParam = str(ws1.cell(row=i, column=5).value)
                TypeParam = str(ws1.cell(row=i, column=4).value)
                SubTypeParam = str(ws1.cell(row=i, column=6).value)
                Inversia = str(ws1.cell(row=i, column=7).value)

                if (NumParam== 'None'):
                    if 'Резерв' in Name:
                        NumParam="0"

                if (TypeParam== 'None'):
                    if 'Резерв' in Name:
                         TypeParam="0"
                elif (TypeParam== 'ДП'):
                     TypeParam="0"
                elif (TypeParam== 'ИП'):
                     TypeParam="1"

                if (SubTypeParam== 'None'):
                     SubTypeParam="0"

                if (Inversia== 'None'):
                      Inversia="0"

            if (maxstr-1)==i:
                self.strArray.append("\n(No:="+NumParam+", PointType:="+TypeParam+", Inversiya:="+Inversia+", SubType:="+SubTypeParam+")]; (*"+"Sig"+str(i-1)+"  "+Name+"*)")
                self.strArray.append("\nEND_VAR")
            else:
                self.strArray.append("\n(No:="+NumParam+", PointType:="+TypeParam+", Inversiya:="+Inversia+", SubType:="+SubTypeParam+"), (*"+"Sig"+str(i-1)+"  "+Name+"*)")
        
        self.savefile()

    def genDS(self):
        wb = load_workbook(self.filename)
        ws1=wb["ДВ"]
        Name=""
        NumMode=""
        NumChannel=""
        maxstr=1000
        self.strArray=[]

        self.strArray.append('VAR_GLOBAL CONSTANT')
        self.strArray.append('\nConfig_DS :  ARRAY[1..1000] OF sConfig_DS:=[')

        for i in range(2,maxstr):
            Name = str(ws1.cell(row=i, column=2).value)

            if Name == 'None':
                Name="Резерв"
                NumMode="0"
                NumChannel = "0"
            else:
                NumMode = str(ws1.cell(row=i, column=4).value)
                NumChannel = str(ws1.cell(row=i, column=5).value)

                if (NumMode== 'None'):
                    if 'Резерв' in Name:
                        NumMode="0"
                if (NumChannel== 'None'):
                    if 'Резерв' in Name:
                        NumChannel="0"

            if (maxstr-1)==i:
                self.strArray.append("\n(NumMod:="+NumMode+", NumChan:="+NumChannel+")]; (*"+"ДВ"+str(i-1)+"  "+Name+"*)")
                self.strArray.append("\nEND_VAR")
            else:
                self.strArray.append("\n(NumMod:="+NumMode+", NumChan:="+NumChannel+"), (*"+"ДВ"+str(i-1)+"  "+Name+"*)")    
        
        self.savefile()

        


    def genDP(self):
        wb = load_workbook(self.filename)
        ws1=wb["ДП"]
        Name=""
        NumMode=""
        NumChannel=""
        color=[0,0]
        maxstr=3000
        self.strArray=[]

        self.strArray.append('VAR_GLOBAL CONSTANT')
        self.strArray.append('\nConfig_DP :  ARRAY[1..3000] OF sConfig_DP:=[')

        for i in range(2,maxstr):
            Name = str(ws1.cell(row=i, column=2).value)

            if Name == 'None':
                Name="Резерв"
                NumMode="0"
                NumChannel = "0"
                color[0]=1
                color[1]=1
            else:


                NumMode = str(ws1.cell(row=i, column=4).value)
                NumChannel = str(ws1.cell(row=i, column=5).value)

                if (NumMode== 'None'):
                    if 'Резерв' in Name:
                        NumMode="0"
                if (NumChannel== 'None'):
                    if 'Резерв' in Name:
                        NumChannel="0"

                for j in range(0,2):
                    if (ws1.cell(row=i, column=7+j).value == "красный"):
                        color[j]=8
                    elif  (ws1.cell(row=i, column=7+j).value == "желтый"):
                        color[j]=4
                    elif  (ws1.cell(row=i, column=7+j).value == "зеленый"):
                        color[j]=2
                    elif  (ws1.cell(row=i, column=7+j).value == "серый"):
                        color[j]=1
                    else:
                        color[j]=1
            if (maxstr-1)==i:
                self.strArray.append("\n(NumMod:="+NumMode+", NumChan:="+NumChannel+", LowColor:="+str(color[0])+", HighColor:="+str(color[1])+")]; (*"+"ДП"+str(i-1)+"  "+Name+"*)")
                self.strArray.append("\nEND_VAR")
            else:
                self.strArray.append("\n(NumMod:="+NumMode+", NumChan:="+NumChannel+", LowColor:="+str(color[0])+", HighColor:="+str(color[1])+"), (*"+"ДП"+str(i-1)+"  "+Name+"*)")    
        
        self.savefile()


    def savefile(self):
        newfile=fd.asksaveasfile(title="Сохранить файл", defaultextension=".txt", filetypes=(("Текстовый файл", "*.txt"),))
        if newfile:
            newfile.writelines(self.strArray)
            newfile.close()


    def main(self):
        self.filename = askopenfilename()
        self.button1.configure(text=self.filename)

    def get_rgb(rgb):
        return "#%02x%02x%02x" % rgb  

    def createWidgets(self):
       self.geometry=root.geometry('380x820')
       self.button1 = Button(root, text=self.filename, width=380, height=5, bg='#FE0000', command=self.main)
       self.button1.pack()
       self.button2 = Button(root, text="Генерировать ДП", width=380, height=5, bg='#00ff00', command=self.genDP)
       self.button2.pack()
       self.button3 = Button(root, text="Генерировать ДВ", width=380, height=5, bg='#00Ef00', command=self.genDS)
       self.button3.pack()
       self.button4 = Button(root, text="Генерировать сигнализации", width=380, height=5, bg='#00Df00', command=self.genSignaliz)
       self.button4.pack()
       self.button5 = Button(root, text="Генерировать задвижки", width=380, height=5, bg='#00Cf00', command=self.genZD)
       self.button5.pack()
       self.button6 = Button(root, text="Генерировать ИП", width=380, height=5, bg='#00Bf00', command=self.genIP)
       self.button6.pack()
       self.button7 = Button(root, text="Генерировать табло и сирены", width=380, height=5, bg='#00Af00', command=self.genUTS)
       self.button7.pack()
       self.button8 = Button(root, text="Генерировать предельные параметры", width=380, height=5, bg='#009f00', command=self.genKTPRS)
       self.button8.pack()
       self.button9 = Button(root, text="Генерировать вспомсистемы", width=380, height=5, bg='#008f00', command=self.genUVS)
       self.button9.pack()
       self.label = Label (root, text="*Генератор работает с xlsx файлами", bg='#007f00',width=380, height=5)
       self.label.pack()
       datetime.datetime.now()


    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.geometry=None
        self.button1 = None
        self.filename = 'Выберете ИО для генерации'
        self.pack()
        self.createWidgets()


root = Tk()
root.title("Генератор инициализации переменных для Прософт")

app = Application(master=root)

app.mainloop()
