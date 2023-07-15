import array
from ast import And
from statistics import geometric_mean
from tkinter import *
from tkinter.filedialog import askopenfilename, asksaveasfile
from turtle import bgcolor, left
import openpyxl
from openpyxl import load_workbook
import tkinter.filedialog as fd

class Application(Frame):

    def genDP(self):
        wb = load_workbook(self.filename)
        ws1=wb["ДП"]
        Name=""
        NumMode=""
        NumChannel=""
        color=[0,0]
        maxstr=3000
        self.strArray=[]

        self.strArray.append('VAR_GLOBAL CONSTANT')
        self.strArray.append('\nConfig_DP :  ARRAY[1..3000] OF sConfig_DP:=[')

        for i in range(2,maxstr):
            Name = str(ws1.cell(row=i, column=2).value)

            if Name == 'None':
                Name="Резерв"
                NumMode="0"
                NumChannel = "0"
                color[0]=1
                color[1]=1
            else:


                NumMode = str(ws1.cell(row=i, column=4).value)
                NumChannel = str(ws1.cell(row=i, column=5).value)

                if (NumMode== 'None'):
                    if 'Резерв' in Name:
                        NumMode="0"
                if (NumChannel== 'None'):
                    if 'Резерв' in Name:
                        NumChannel="0"

                for j in range(0,2):
                    if (ws1.cell(row=i, column=7+j).value == "красный"):
                        color[j]=8
                    elif  (ws1.cell(row=i, column=7+j).value == "желтый"):
                        color[j]=4
                    elif  (ws1.cell(row=i, column=7+j).value == "зеленый"):
                        color[j]=2
                    elif  (ws1.cell(row=i, column=7+j).value == "серый"):
                        color[j]=1
                    else:
                        color[j]=1
            if (maxstr-1)==i:
                self.strArray.append("\n(NumMod:="+NumMode+", NumChan:="+NumChannel+", LowColor:="+str(color[0])+", HighColor:="+str(color[1])+")]; (*"+"ДП"+str(i-1)+"  "+Name+"*)")
                self.strArray.append("\nEND_VAR")
            else:
                self.strArray.append("\n(NumMod:="+NumMode+", NumChan:="+NumChannel+", LowColor:="+str(color[0])+", HighColor:="+str(color[1])+"), (*"+"ДП"+str(i-1)+"  "+Name+"*)")    
        
        self.savefile()


    def savefile(self):
        newfile=fd.asksaveasfile(title="Сохранить файл", defaultextension=".txt", filetypes=(("Текстовый файл", "*.txt"),))
        if newfile:
            newfile.writelines(self.strArray)
            newfile.close()


    def main(self):
        self.filename = askopenfilename()
        self.button1.configure(text=self.filename)


    def createWidgets(self):
       self.geometry=root.geometry('380x500')
       self.button1 = Button(root, text=self.filename, width=380, height=5, bg='lightgreen', command=self.main)
       self.button1.pack()
       self.button2 = Button(root, text="Gener DP", width=380, height=5, bg='lightyellow', command=self.genDP)
       self.button2.pack()

    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.geometry=None
        self.button1 = None
        self.filename = 'No choose file'
        self.pack()
        self.createWidgets()


root = Tk()
root.title("Generator initialization of variable for Prosoft")
#root.geometry('380x500')

app = Application(master=root)

app.mainloop()
